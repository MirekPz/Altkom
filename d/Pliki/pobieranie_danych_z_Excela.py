from openpyxl import load_workbook

workbook = load_workbook('wynik.xlsx')
print(workbook)
print(dir(workbook))

print(workbook, workbook.sheetnames)
print(workbook['wynik'])

karta = workbook['wynik']
print([x.value for x in karta[2]])
print(karta['a2'])
print(karta['a2'].value)

for i in range(1, 11):
    print([x.value for x in karta[i]])
    print(karta[i][1].value)
