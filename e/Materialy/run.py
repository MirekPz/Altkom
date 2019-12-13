from openpyxl import load_workbook
wb2 = load_workbook('test.xlsx')
print(wb2.sheetnames)
print(dir(wb2))
print(dir(wb2.worksheets[0]))
worksheet = wb2[wb2.worksheets[0].title]

print(worksheet["a2"].value)

first_row = worksheet[2]
print(first_row)
for cell in first_row:
     print(cell.value)
